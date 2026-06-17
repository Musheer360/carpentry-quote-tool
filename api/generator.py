"""
Excel generator for the Carpentry Quote Tool.

Produces ONE workbook per client. Inside:
  - "الأسعار" (Prices): the ONLY place unit prices are typed. Every item sheet
    references these cells, so updating a price updates the whole quote.
  - one sheet PER ITEM: Arabic, right-to-left, SAR, beautifully formatted.
    Every money cell is a real Excel formula (clickable):
      * line total      = qty * price-reference
      * subtotal        = SUM(...)
      * finishing area  = length * width * count
      * labour transport= work-days * 2 * rate
      * category totals  = SUMIF(...)
      * % of total       = category / grand-total
    A "تفاصيل الحسابات" (Detailed Calculations) section restates every line as a
    live concatenation formula -> "7 × 150.00 = 1,050.00" that recalculates.
  - "الملخص" (Summary): per-item totals, grand total, and an estimated timeline.

Nothing computed is ever hardcoded.
"""

from __future__ import annotations

import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

from estimator import estimate as estimate_unit

# ---------------------------------------------------------------- styling ---
SAR_FMT = '#,##0.00\\ "ر.س"'
QTY_FMT = '#,##0.###'
PCT_FMT = '0.0%'

C_BRAND = "1F4E5F"      # deep teal (banners / headers)
C_BRAND2 = "2E7D92"     # lighter teal (section headers)
C_HEADERFG = "FFFFFF"
C_TOTAL = "FCE4B6"      # warm gold (totals)
C_ZEBRA = "F4F8FA"      # very light blue (zebra)
C_META = "EDF3F5"
C_BREAK = "EAF3E7"
C_NOTE = "FFF6E5"
C_TABCLR = "1F4E5F"

_thin = Side(style="thin", color="C7D2D8")
_med = Side(style="medium", color=C_BRAND)
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOTAL = Border(left=_thin, right=_thin, top=_med, bottom=_med)

FONT = "Arial"
F_BANNER = Font(name=FONT, size=16, bold=True, color=C_HEADERFG)
F_SUBBAN = Font(name=FONT, size=11, color=C_HEADERFG)
F_SECTION = Font(name=FONT, size=12, bold=True, color=C_HEADERFG)
F_HEAD = Font(name=FONT, size=11, bold=True, color=C_HEADERFG)
F_LABEL = Font(name=FONT, size=10, bold=True, color="33474F")
F_CELL = Font(name=FONT, size=11, color="1F2D33")
F_TOTAL = Font(name=FONT, size=12, bold=True, color="1F2D33")
F_NOTE = Font(name=FONT, size=10, italic=True, color="B00000")
F_MONO = Font(name="Consolas", size=10, color="1F2D33")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# table columns (RTL display flips these visually):  J=Total K=Price L=Qty M=Material
COL_TOTAL, COL_PRICE, COL_QTY, COL_NAME = 10, 11, 12, 13
COL_CAT = 15            # hidden category tag
COL_L, COL_W, COL_CNT = 16, 17, 18  # hidden dim helpers for finishing area


def _safe_sheet_title(name, used):
    bad = '[]:*?/\\'
    clean = "".join("_" if ch in bad else ch for ch in (name or "بند")).strip() or "بند"
    clean = clean[:31]
    base, i = clean, 2
    while clean.lower() in used:
        sfx = f" ({i})"
        clean = base[: 31 - len(sfx)] + sfx
        i += 1
    used.add(clean.lower())
    return clean


class QuoteGenerator:
    def __init__(self, pricebook):
        self.pb = pricebook
        self.items_by_id = {it["id"]: it for it in pricebook.get("items", [])}
        self.cat_by_id = {c["id"]: c for c in pricebook.get("categories", [])}
        self.week_days = pricebook.get("week_days", 6) or 6
        self.month_days = pricebook.get("month_days", 26) or 26

    # -- price resolution ----------------------------------------------------
    def _price(self, item_id, overrides):
        if overrides and item_id in overrides and overrides[item_id] not in (None, ""):
            try:
                return float(overrides[item_id])
            except (TypeError, ValueError):
                pass
        it = self.items_by_id.get(item_id)
        return float(it["price"]) if it else 0.0

    # -- public --------------------------------------------------------------
    def build(self, project, data_dir, out_path):
        wb = Workbook()
        wb.remove(wb.active)
        overrides = project.get("price_overrides", {}) or {}
        items = project.get("items", []) or []

        used_ids, seen = [], set()
        for item in items:
            for pid in self._ids_in_item(item):
                if pid and pid not in seen and pid in self.items_by_id:
                    seen.add(pid)
                    used_ids.append(pid)

        price_ref = self._build_prices_sheet(wb, used_ids, overrides)

        used_titles = set()
        item_sheets = []
        for item in items:
            title = _safe_sheet_title(item.get("name_ar") or item.get("name_en"), used_titles)
            ws = wb.create_sheet(title=title)
            ws.sheet_properties.tabColor = C_TABCLR
            if item.get("kind") == "unit":
                total_cell, days_cell = self._build_unit_sheet(ws, project, item)
            else:
                total_cell, days_cell = self._build_item_sheet(ws, project, item, price_ref, data_dir)
            item_sheets.append((title, total_cell, days_cell))

        self._build_summary_sheet(wb, project, item_sheets)
        wb.move_sheet(wb["الملخص"], -(len(wb.sheetnames) - 1))

        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        wb.save(out_path)
        return out_path

    # -- id discovery --------------------------------------------------------
    def _ids_in_item(self, item):
        ids = []
        if item.get("kind") == "unit":
            return ids
        for ln in item.get("lines", []) or []:
            if ln.get("item_id"):
                ids.append(ln["item_id"])
            if ln.get("finish") == "polish":
                ids.append("polish_wood")
            elif ln.get("finish") == "paint":
                ids.append("paint_wood")
            if ln.get("process") == "cnc":
                ids.append("cnc_cut")
            elif ln.get("process") == "pvc_edge":
                ids.append("edge_pvc")
        if float(item.get("labour_days") or 0) > 0:
            ids += ["labour_day", "transport_labour"]
        if float(item.get("material_transport_trips") or 0) > 0:
            ids.append("transport_mat")
        return ids

    # -- Prices sheet --------------------------------------------------------
    def _build_prices_sheet(self, wb, used_ids, overrides):
        ws = wb.create_sheet(title="الأسعار")
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = "B59410"
        widths = {"A": 3, "B": 24, "C": 42, "D": 12, "E": 16}
        for c, w in widths.items():
            ws.column_dimensions[c].width = w

        ws.merge_cells("B1:E1")
        b = ws["B1"]
        b.value = "قائمة الأسعار  —  المصدر الوحيد لجميع الأسعار"
        b.font = F_BANNER
        b.alignment = CENTER
        b.fill = PatternFill("solid", fgColor=C_BRAND)
        ws.row_dimensions[1].height = 30

        for col, text in zip(range(2, 6), ["التصنيف", "المادة / الخدمة", "الوحدة", "السعر (ر.س)"]):
            c = ws.cell(row=2, column=col, value=text)
            c.font = F_HEAD
            c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2)
            c.border = BORDER

        ref = {}
        r = 3
        for pid in used_ids:
            it = self.items_by_id[pid]
            cat = self.cat_by_id.get(it.get("category"), {})
            ws.cell(row=r, column=2, value=cat.get("name_ar", "")).font = F_CELL
            ws.cell(row=r, column=3, value=it.get("name_ar", "")).font = F_CELL
            ws.cell(row=r, column=4, value=it.get("unit", "")).font = F_CELL
            pc = ws.cell(row=r, column=5, value=self._price(pid, overrides))
            pc.font = Font(name=FONT, size=11, bold=True)
            pc.number_format = SAR_FMT
            fill = PatternFill("solid", fgColor=C_ZEBRA) if r % 2 else None
            for col in range(2, 6):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER
                cell.alignment = RIGHT if col < 5 else CENTER
                if fill:
                    cell.fill = fill
            ref[pid] = f"'الأسعار'!$E${r}"
            r += 1
        ws.freeze_panes = "A3"
        return ref

    # -- one item sheet ------------------------------------------------------
    def _build_item_sheet(self, ws, project, item, price_ref, data_dir):
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        for col, w in {"A": 2, "B": 13, "C": 13, "D": 13, "E": 13, "F": 4,
                       "G": 4, "H": 4, "I": 4, "J": 18, "K": 14, "L": 11, "M": 40}.items():
            ws.column_dimensions[col].width = w

        # ---- banner ----
        ws.merge_cells("B1:M1")
        ban = ws["B1"]
        ban.value = project.get("company_name_ar") or project.get("company_name_en") or "عرض سعر"
        ban.font = F_BANNER
        ban.alignment = CENTER
        ban.fill = PatternFill("solid", fgColor=C_BRAND)
        ws.row_dimensions[1].height = 30
        ws.merge_cells("B2:M2")
        sub = ws["B2"]
        sub.value = "عرض سعر أعمال النجارة و الديكور الداخلي"
        sub.font = F_SUBBAN
        sub.alignment = CENTER
        sub.fill = PatternFill("solid", fgColor=C_BRAND2)

        # ---- metadata (right J:M) ----
        meta = [
            ("اسم العميل", project.get("client_name_ar") or project.get("client_name_en") or ""),
            ("اسم الوحدة / المشروع", project.get("unit_name") or ""),
            ("الموقع / المكان", item.get("place_ar") or item.get("place_en") or project.get("location") or ""),
            ("البند", item.get("name_ar") or item.get("name_en") or ""),
        ]
        row = 3
        for label, value in meta:
            lc = ws.cell(row=row, column=COL_NAME, value=label)
            lc.font = F_LABEL
            lc.alignment = RIGHT
            lc.fill = PatternFill("solid", fgColor=C_META)
            lc.border = BORDER
            ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
            vc = ws.cell(row=row, column=COL_TOTAL, value=value)
            vc.font = F_CELL
            vc.alignment = RIGHT
            vc.border = BORDER
            row += 1
        # date + time
        for label, value, fmt in [("التاريخ", "=TODAY()", "yyyy-mm-dd"),
                                  ("الوقت", item.get("time") or "", None)]:
            lc = ws.cell(row=row, column=COL_NAME, value=label)
            lc.font = F_LABEL
            lc.alignment = RIGHT
            lc.fill = PatternFill("solid", fgColor=C_META)
            lc.border = BORDER
            ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
            vc = ws.cell(row=row, column=COL_TOTAL, value=value)
            if fmt:
                vc.number_format = fmt
            vc.font = F_CELL
            vc.alignment = RIGHT
            vc.border = BORDER
            row += 1
        meta_bottom = row - 1

        # image on the left (B3 area)
        self._place_image(ws, item, data_dir, anchor="B3")

        # ---- materials table ----
        ws.column_dimensions[get_column_letter(COL_CAT)].hidden = True
        for c in (COL_L, COL_W, COL_CNT):
            ws.column_dimensions[get_column_letter(c)].hidden = True

        start = max(meta_bottom, 8) + 2
        self._section(ws, start, "المواد و الأعمال")
        hdr = start + 1
        for col, text in {COL_TOTAL: "قيمة الإجمالي", COL_PRICE: "السعر",
                          COL_QTY: "العدد", COL_NAME: "مواد العمل و الإكسسوارات"}.items():
            c = ws.cell(row=hdr, column=col, value=text)
            c.font = F_HEAD
            c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2)
            c.border = BORDER

        first_data = hdr + 1
        r = first_data
        line_rows = []   # (row, kind, line) for the detailed-calc section
        for line in item.get("lines", []) or []:
            r, written = self._write_line(ws, r, line, price_ref)
            line_rows += written
        r, lt_rows = self._write_labour_transport(ws, r, item, price_ref)
        line_rows += lt_rows
        last_data = max(r - 1, first_data)

        # zebra striping
        for rr in range(first_data, last_data + 1):
            if (rr - first_data) % 2 == 1:
                for col in range(COL_TOTAL, COL_NAME + 1):
                    cell = ws.cell(row=rr, column=col)
                    if cell.fill.fgColor.rgb in (None, "00000000"):
                        cell.fill = PatternFill("solid", fgColor=C_ZEBRA)

        # subtotal
        total_row = r
        tc = ws.cell(row=total_row, column=COL_TOTAL, value=f"=SUM(J{first_data}:J{last_data})")
        tc.font = F_TOTAL
        tc.number_format = SAR_FMT
        tc.alignment = CENTER
        tc.fill = PatternFill("solid", fgColor=C_TOTAL)
        tc.border = BORDER_TOTAL
        lbl = ws.cell(row=total_row, column=COL_NAME, value="إجمالي تكلفة البند")
        lbl.font = F_TOTAL
        lbl.alignment = RIGHT
        lbl.fill = PatternFill("solid", fgColor=C_TOTAL)
        for col in (COL_PRICE, COL_QTY):
            cc = ws.cell(row=total_row, column=col)
            cc.fill = PatternFill("solid", fgColor=C_TOTAL)
            cc.border = BORDER_TOTAL
        lbl.border = BORDER_TOTAL
        total_cell = f"'{ws.title}'!$J${total_row}"

        r = total_row + 1
        if item.get("note_ar"):
            ws.merge_cells(start_row=r, start_column=COL_TOTAL, end_row=r, end_column=COL_NAME)
            nc = ws.cell(row=r, column=COL_TOTAL, value="ملاحظة - " + item["note_ar"])
            nc.font = F_NOTE
            nc.alignment = RIGHT
            nc.fill = PatternFill("solid", fgColor=C_NOTE)
            r += 1

        # ---- detailed calculations (live concat formulas) ----
        r = self._write_calc_details(ws, r + 1, line_rows, total_row)

        # ---- category breakdown with % ----
        r = self._write_breakdown(ws, r + 1, first_data, last_data, total_row)

        # ---- time estimate ----
        days_cell = self._write_time(ws, r + 1, item)

        ws.freeze_panes = f"A{first_data}"
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        return total_cell, days_cell

    # -- section header helper ----------------------------------------------
    def _section(self, ws, row, text):
        ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_NAME)
        c = ws.cell(row=row, column=COL_TOTAL, value=text)
        c.font = F_SECTION
        c.alignment = RIGHT
        c.fill = PatternFill("solid", fgColor=C_BRAND)
        ws.row_dimensions[row].height = 22
        return row

    # -- write one input line -> sheet rows ----------------------------------
    def _write_line(self, ws, r, line, price_ref):
        kind = line.get("kind", "simple")
        pid = line.get("item_id")
        qty = self._num(line.get("qty"))
        written = []

        if pid:
            self._material_row(ws, r, pid, price_ref, qty_value=qty)
            written.append((r, "simple", {}))
            r += 1

        if kind == "mdf" and line.get("finish") in ("polish", "paint"):
            finish_id = "polish_wood" if line["finish"] == "polish" else "paint_wood"
            length = self._num(line.get("length"))
            width = self._num(line.get("width"))
            count = self._num(line.get("finish_count"), default=qty or 1)
            ws.cell(row=r, column=COL_L, value=length)
            ws.cell(row=r, column=COL_W, value=width)
            ws.cell(row=r, column=COL_CNT, value=count)
            self._material_row(ws, r, finish_id, price_ref, qty_formula=f"=P{r}*Q{r}*R{r}")
            written.append((r, "area", {}))
            r += 1

        if kind == "laminated" and line.get("process") in ("cnc", "pvc_edge"):
            if line["process"] == "cnc":
                proc_id, proc_qty = "cnc_cut", self._num(line.get("process_qty"), default=qty or 1)
            else:
                proc_id, proc_qty = "edge_pvc", self._num(line.get("process_qty"))
            self._material_row(ws, r, proc_id, price_ref, qty_value=proc_qty)
            written.append((r, "simple", {}))
            r += 1
        return r, written

    def _material_row(self, ws, r, pid, price_ref, qty_value=None, qty_formula=None):
        it = self.items_by_id.get(pid, {})
        nc = ws.cell(row=r, column=COL_NAME, value=it.get("name_ar", pid))
        nc.font = F_CELL
        nc.alignment = RIGHT
        lc = ws.cell(row=r, column=COL_QTY, value=qty_formula if qty_formula else qty_value)
        lc.number_format = QTY_FMT
        lc.alignment = CENTER
        lc.font = F_CELL
        kref = price_ref.get(pid)
        kc = ws.cell(row=r, column=COL_PRICE, value=(f"={kref}" if kref else 0))
        kc.number_format = SAR_FMT
        kc.alignment = CENTER
        kc.font = F_CELL
        jc = ws.cell(row=r, column=COL_TOTAL, value=f"=L{r}*K{r}")
        jc.number_format = SAR_FMT
        jc.alignment = CENTER
        jc.font = F_CELL
        ws.cell(row=r, column=COL_CAT, value=it.get("category", ""))
        for col in range(COL_TOTAL, COL_NAME + 1):
            ws.cell(row=r, column=col).border = BORDER

    def _write_labour_transport(self, ws, r, item, price_ref):
        days = self._num(item.get("labour_days"))
        trips = self._num(item.get("material_transport_trips"))
        written = []
        if days > 0:
            self._material_row(ws, r, "labour_day", price_ref, qty_value=days)
            day_row = r
            written.append((r, "simple", {}))
            r += 1
            self._material_row(ws, r, "transport_labour", price_ref, qty_formula=f"=L{day_row}*2")
            written.append((r, "limo", {"day_row": day_row}))
            r += 1
        if trips > 0:
            self._material_row(ws, r, "transport_mat", price_ref, qty_value=trips)
            written.append((r, "simple", {}))
            r += 1
        return r, written

    # -- detailed calculations (concatenation formulas) ----------------------
    def _write_calc_details(self, ws, row, line_rows, total_row):
        self._section(ws, row, "تفاصيل الحسابات  (العدد × السعر = الإجمالي)")
        row += 1
        # header
        hc = ws.cell(row=row, column=COL_NAME, value="البند")
        hc.font = F_HEAD
        hc.alignment = CENTER
        hc.fill = PatternFill("solid", fgColor=C_BRAND2)
        hc.border = BORDER
        ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
        ec = ws.cell(row=row, column=COL_TOTAL, value="المعادلة و النتيجة")
        ec.font = F_HEAD
        ec.alignment = CENTER
        ec.fill = PatternFill("solid", fgColor=C_BRAND2)
        ec.border = BORDER
        for col in (COL_PRICE, COL_QTY):
            ws.cell(row=row, column=col).border = BORDER
        row += 1

        for i, (lr, ltype, meta) in enumerate(line_rows):
            # name mirrors the line's material name
            nc = ws.cell(row=row, column=COL_NAME, value=f"=M{lr}")
            nc.font = F_CELL
            nc.alignment = RIGHT
            nc.border = BORDER
            # build a live concatenation formula that reads back qty × price = total
            if ltype == "area":
                expr = (f'=TEXT(P{lr},"0.##")&" × "&TEXT(Q{lr},"0.##")&" × "&TEXT(R{lr},"0.##")'
                        f'&" م² × "&TEXT(K{lr},"#,##0.00")&" = "&TEXT(J{lr},"#,##0.00")&" ر.س"')
            elif ltype == "limo":
                dr = meta["day_row"]
                expr = (f'="("&TEXT(L{dr},"0.##")&" يوم × 2) × "&TEXT(K{lr},"#,##0.00")'
                        f'&" = "&TEXT(J{lr},"#,##0.00")&" ر.س"')
            else:
                expr = (f'=TEXT(L{lr},"0.###")&" × "&TEXT(K{lr},"#,##0.00")'
                        f'&" = "&TEXT(J{lr},"#,##0.00")&" ر.س"')
            ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
            xc = ws.cell(row=row, column=COL_TOTAL, value=expr)
            xc.font = F_MONO
            xc.alignment = LEFT
            xc.border = BORDER
            if i % 2 == 1:
                for col in range(COL_TOTAL, COL_NAME + 1):
                    ws.cell(row=row, column=col).fill = PatternFill("solid", fgColor=C_ZEBRA)
            ws.cell(row=row, column=COL_PRICE).border = BORDER
            ws.cell(row=row, column=COL_QTY).border = BORDER
            row += 1

        # grand line: sum reference
        ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
        gc = ws.cell(row=row, column=COL_TOTAL,
                     value=f'="مجموع البنود = "&TEXT(J{total_row},"#,##0.00")&" ر.س"')
        gc.font = F_TOTAL
        gc.alignment = LEFT
        gc.fill = PatternFill("solid", fgColor=C_TOTAL)
        gc.border = BORDER_TOTAL
        gl = ws.cell(row=row, column=COL_NAME, value="الإجمالي")
        gl.font = F_TOTAL
        gl.alignment = RIGHT
        gl.fill = PatternFill("solid", fgColor=C_TOTAL)
        gl.border = BORDER_TOTAL
        return row + 1

    # -- category breakdown with % -------------------------------------------
    def _write_breakdown(self, ws, row, first_data, last_data, total_row):
        present = set()
        for rr in range(first_data, last_data + 1):
            v = ws.cell(row=rr, column=COL_CAT).value
            if v:
                present.add(v)
        cats = [c for c in self.pb.get("categories", []) if c["id"] in present]
        if not cats:
            return row
        self._section(ws, row, "تفصيل التكلفة حسب التصنيف")
        row += 1
        for col, text in {COL_TOTAL: "النسبة %", COL_PRICE: "القيمة", COL_NAME: "التصنيف"}.items():
            c = ws.cell(row=row, column=col, value=text)
            c.font = F_HEAD
            c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2)
            c.border = BORDER
        ws.merge_cells(start_row=row, start_column=COL_PRICE, end_row=row, end_column=COL_QTY)
        row += 1
        for i, cat in enumerate(cats):
            sumif = (f'=SUMIF($O${first_data}:$O${last_data},"{cat["id"]}",'
                     f'$J${first_data}:$J${last_data})')
            lc = ws.cell(row=row, column=COL_NAME, value=cat["name_ar"])
            lc.font = F_CELL
            lc.alignment = RIGHT
            ws.merge_cells(start_row=row, start_column=COL_PRICE, end_row=row, end_column=COL_QTY)
            vc = ws.cell(row=row, column=COL_PRICE, value=sumif)
            vc.number_format = SAR_FMT
            vc.alignment = CENTER
            vc.font = F_CELL
            pc = ws.cell(row=row, column=COL_TOTAL, value=f"=IF($J${total_row}=0,0,K{row}/$J${total_row})")
            pc.number_format = PCT_FMT
            pc.alignment = CENTER
            pc.font = F_CELL
            fill = PatternFill("solid", fgColor=C_BREAK) if i % 2 == 0 else None
            for col in (COL_TOTAL, COL_PRICE, COL_QTY, COL_NAME):
                cell = ws.cell(row=row, column=col)
                cell.border = BORDER
                if fill:
                    cell.fill = fill
            row += 1
        return row

    # -- time estimate -------------------------------------------------------
    def _write_time(self, ws, row, item):
        days = self._num(item.get("labour_days"))
        self._section(ws, row, "الوقت التقديري للتنفيذ")
        row += 1
        base = row
        rows = [
            ("عدد أيام العمل", days, QTY_FMT),
            (f"بالأسابيع (أسبوع = {self.week_days} يوم)", f"=L{base}/{self.week_days}", "#,##0.0"),
            (f"بالأشهر (شهر = {self.month_days} يوم)", f"=L{base}/{self.month_days}", "#,##0.0"),
        ]
        days_cell = f"'{ws.title}'!$L${base}"
        for i, (label, val, fmt) in enumerate(rows):
            rr = base + i
            lc = ws.cell(row=rr, column=COL_NAME, value=label)
            lc.font = F_CELL
            lc.alignment = RIGHT
            ws.merge_cells(start_row=rr, start_column=COL_TOTAL, end_row=rr, end_column=COL_QTY)
            vc = ws.cell(row=rr, column=COL_TOTAL, value=val)
            vc.number_format = fmt
            vc.alignment = CENTER
            vc.font = F_CELL
            for col in (COL_TOTAL, COL_NAME):
                ws.cell(row=rr, column=col).border = BORDER
        return days_cell

    # -- Smart Unit sheet ----------------------------------------------------
    def _build_unit_sheet(self, ws, project, item):
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        for col, w in {"A": 2, "B": 13, "C": 13, "D": 13, "E": 13, "F": 4, "G": 4,
                       "H": 4, "I": 4, "J": 18, "K": 14, "L": 11, "M": 40}.items():
            ws.column_dimensions[col].width = w

        est = estimate_unit(item.get("unit", {}), self.pb)
        u = item.get("unit", {})

        # banner
        ws.merge_cells("B1:M1")
        ban = ws["B1"]
        ban.value = project.get("company_name_ar") or project.get("company_name_en") or "عرض سعر"
        ban.font = F_BANNER; ban.alignment = CENTER
        ban.fill = PatternFill("solid", fgColor=C_BRAND)
        ws.row_dimensions[1].height = 30
        ws.merge_cells("B2:M2")
        sub = ws["B2"]
        type_ar = {"wardrobe": "خزانة", "tv_wall": "جدار تلفاز", "wall_paneling": "تلبيس جدار"}.get(u.get("unit_type"), "وحدة")
        sub.value = "وحدة ذكية - %s" % type_ar
        sub.font = F_SUBBAN; sub.alignment = CENTER
        sub.fill = PatternFill("solid", fgColor=C_BRAND2)

        meta = [
            ("اسم العميل", project.get("client_name_ar") or project.get("client_name_en") or ""),
            ("البند", item.get("name_ar") or item.get("name_en") or ""),
            ("المقاسات (سم) ع×ط×ع", "%s × %s × %s" % (u.get("width_cm"), u.get("height_cm"), u.get("depth_cm"))),
            ("المكان", item.get("place_ar") or ""),
        ]
        row = 3
        for label, value in meta:
            lc = ws.cell(row=row, column=COL_NAME, value=label)
            lc.font = F_LABEL; lc.alignment = RIGHT
            lc.fill = PatternFill("solid", fgColor=C_META); lc.border = BORDER
            ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
            vc = ws.cell(row=row, column=COL_TOTAL, value=value)
            vc.font = F_CELL; vc.alignment = RIGHT; vc.border = BORDER
            row += 1
        ws.cell(row=row, column=COL_NAME, value="التاريخ").font = F_LABEL
        ws.cell(row=row, column=COL_NAME).alignment = RIGHT
        ws.cell(row=row, column=COL_NAME).fill = PatternFill("solid", fgColor=C_META)
        ws.cell(row=row, column=COL_NAME).border = BORDER
        ws.merge_cells(start_row=row, start_column=COL_TOTAL, end_row=row, end_column=COL_QTY)
        dc = ws.cell(row=row, column=COL_TOTAL, value="=TODAY()")
        dc.number_format = "yyyy-mm-dd"; dc.alignment = RIGHT; dc.border = BORDER
        row += 2

        # BOM table (formula-driven totals)
        self._section(ws, row, "قائمة المواد (BOM)")
        row += 1
        for col, text in {COL_TOTAL: "قيمة الإجمالي", COL_PRICE: "السعر", COL_QTY: "العدد", COL_NAME: "البند"}.items():
            c = ws.cell(row=row, column=col, value=text)
            c.font = F_HEAD; c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2); c.border = BORDER
        ws.column_dimensions[get_column_letter(COL_CAT)].hidden = True
        first = row + 1
        r = first
        for b in est["bom"]:
            ws.cell(row=r, column=COL_NAME, value=b["label_ar"] or b["label_en"]).alignment = RIGHT
            ws.cell(row=r, column=COL_NAME).font = F_CELL
            qc = ws.cell(row=r, column=COL_QTY, value=b["qty"]); qc.number_format = QTY_FMT; qc.alignment = CENTER; qc.font = F_CELL
            kc = ws.cell(row=r, column=COL_PRICE, value=b["price"]); kc.number_format = SAR_FMT; kc.alignment = CENTER; kc.font = F_CELL
            jc = ws.cell(row=r, column=COL_TOTAL, value=f"=L{r}*K{r}"); jc.number_format = SAR_FMT; jc.alignment = CENTER; jc.font = F_CELL
            ws.cell(row=r, column=COL_CAT, value=b["cat"])
            for col in range(COL_TOTAL, COL_NAME + 1):
                ws.cell(row=r, column=col).border = BORDER
                if (r - first) % 2 == 1:
                    ws.cell(row=r, column=col).fill = PatternFill("solid", fgColor=C_ZEBRA)
            r += 1
        last = r - 1
        total_row = r
        tc = ws.cell(row=total_row, column=COL_TOTAL, value=f"=SUM(J{first}:J{last})")
        tc.font = F_TOTAL; tc.number_format = SAR_FMT; tc.alignment = CENTER
        tc.fill = PatternFill("solid", fgColor=C_TOTAL); tc.border = BORDER_TOTAL
        lbl = ws.cell(row=total_row, column=COL_NAME, value="الإجمالي الكلي للوحدة")
        lbl.font = F_TOTAL; lbl.alignment = RIGHT
        lbl.fill = PatternFill("solid", fgColor=C_TOTAL); lbl.border = BORDER_TOTAL
        for col in (COL_PRICE, COL_QTY):
            ws.cell(row=total_row, column=col).fill = PatternFill("solid", fgColor=C_TOTAL)
            ws.cell(row=total_row, column=col).border = BORDER_TOTAL
        total_cell = f"'{ws.title}'!$J${total_row}"

        # cost breakdown by category (SUMIF)
        r = total_row + 2
        self._section(ws, r, "تفصيل التكلفة حسب التصنيف")
        r += 1
        cats = [c for c in self.pb.get("categories", [])]
        present = set(b["cat"] for b in est["bom"])
        for cat in cats:
            if cat["id"] not in present:
                continue
            lc = ws.cell(row=r, column=COL_NAME, value=cat["name_ar"]); lc.font = F_CELL; lc.alignment = RIGHT
            ws.merge_cells(start_row=r, start_column=COL_PRICE, end_row=r, end_column=COL_QTY)
            vc = ws.cell(row=r, column=COL_PRICE,
                         value=f'=SUMIF($O${first}:$O${last},"{cat["id"]}",$J${first}:$J${last})')
            vc.number_format = SAR_FMT; vc.alignment = CENTER; vc.font = F_CELL
            pc = ws.cell(row=r, column=COL_TOTAL, value=f"=IF($J${total_row}=0,0,K{r}/$J${total_row})")
            pc.number_format = PCT_FMT; pc.alignment = CENTER; pc.font = F_CELL
            for col in (COL_TOTAL, COL_PRICE, COL_QTY, COL_NAME):
                ws.cell(row=r, column=col).border = BORDER
            r += 1

        # nesting summary
        r += 1
        self._section(ws, r, "ملخص التعشيش (الألواح)")
        r += 1
        for col, text in {COL_TOTAL: "نسبة الاستغلال", COL_PRICE: "عدد الألواح", COL_NAME: "المادة"}.items():
            c = ws.cell(row=r, column=col, value=text); c.font = F_HEAD; c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2); c.border = BORDER
        ws.merge_cells(start_row=r, start_column=COL_PRICE, end_row=r, end_column=COL_QTY)
        r += 1
        for mat, nz in est["nesting"].items():
            ws.cell(row=r, column=COL_NAME, value=nz["material_ar"]).alignment = RIGHT
            ws.cell(row=r, column=COL_NAME).font = F_CELL
            ws.merge_cells(start_row=r, start_column=COL_PRICE, end_row=r, end_column=COL_QTY)
            sc = ws.cell(row=r, column=COL_PRICE, value=nz["sheet_count"]); sc.alignment = CENTER; sc.font = F_CELL
            uc = ws.cell(row=r, column=COL_TOTAL, value=nz["utilization"]); uc.number_format = PCT_FMT; uc.alignment = CENTER; uc.font = F_CELL
            for col in (COL_TOTAL, COL_PRICE, COL_QTY, COL_NAME):
                ws.cell(row=r, column=col).border = BORDER
            r += 1

        # time estimate
        days = float(u.get("labour_days") or 0)
        r += 1
        days_cell = self._write_time(ws, r, {"labour_days": days})

        ws.freeze_panes = f"A{first}"
        ws.print_options.horizontalCentered = True
        ws.page_setup.orientation = "portrait"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        return total_cell, days_cell

    # -- summary -------------------------------------------------------------
    def _build_summary_sheet(self, wb, project, item_sheets):
        ws = wb.create_sheet(title="الملخص")
        ws.sheet_view.rightToLeft = True
        ws.sheet_view.showGridLines = False
        ws.sheet_properties.tabColor = C_TABCLR
        for col, w in {"A": 2, "B": 6, "C": 44, "D": 18, "E": 14}.items():
            ws.column_dimensions[col].width = w

        ws.merge_cells("B1:E1")
        b = ws["B1"]
        b.value = project.get("company_name_ar") or "ملخص المشروع"
        b.font = F_BANNER
        b.alignment = CENTER
        b.fill = PatternFill("solid", fgColor=C_BRAND)
        ws.row_dimensions[1].height = 30
        ws.merge_cells("B2:E2")
        s = ws["B2"]
        s.value = "ملخص المشروع و الإجمالي العام"
        s.font = F_SUBBAN
        s.alignment = CENTER
        s.fill = PatternFill("solid", fgColor=C_BRAND2)

        ws.merge_cells("C3:E3")
        ws["C3"] = "العميل: " + (project.get("client_name_ar") or project.get("client_name_en") or "")
        ws["C3"].font = F_LABEL
        ws["C3"].alignment = RIGHT
        ws["B4"] = "التاريخ"
        ws["B4"].font = F_LABEL
        ws.merge_cells("C4:D4")
        dc = ws["C4"]
        dc.value = "=TODAY()"
        dc.number_format = "yyyy-mm-dd"
        dc.alignment = RIGHT

        hr = 6
        for col, text in {2: "#", 3: "البند", 4: "إجمالي التكلفة", 5: "أيام العمل"}.items():
            c = ws.cell(row=hr, column=col, value=text)
            c.font = F_HEAD
            c.alignment = CENTER
            c.fill = PatternFill("solid", fgColor=C_BRAND2)
            c.border = BORDER

        r = hr + 1
        first = r
        for idx, (title, total_cell, days_cell) in enumerate(item_sheets, start=1):
            ws.cell(row=r, column=2, value=idx).alignment = CENTER
            nc = ws.cell(row=r, column=3, value=f"='{title}'!$M$6")  # item name from sheet
            # fall back to title text if reference is awkward; safer to use literal:
            nc.value = title
            nc.alignment = RIGHT
            nc.font = F_CELL
            tc = ws.cell(row=r, column=4, value=f"={total_cell}")
            tc.number_format = SAR_FMT
            tc.alignment = CENTER
            tc.font = F_CELL
            dc = ws.cell(row=r, column=5, value=f"={days_cell}" if days_cell else 0)
            dc.number_format = QTY_FMT
            dc.alignment = CENTER
            dc.font = F_CELL
            fill = PatternFill("solid", fgColor=C_ZEBRA) if idx % 2 else None
            for col in range(2, 6):
                cell = ws.cell(row=r, column=col)
                cell.border = BORDER
                if fill:
                    cell.fill = fill
            r += 1
        last = r - 1 if r > first else first

        gl = ws.cell(row=r, column=3, value="الإجمالي الكلي للمشروع")
        gl.font = F_TOTAL
        gl.alignment = RIGHT
        gl.fill = PatternFill("solid", fgColor=C_TOTAL)
        gl.border = BORDER_TOTAL
        gc = ws.cell(row=r, column=4, value=f"=SUM(D{first}:D{last})")
        gc.font = F_TOTAL
        gc.number_format = SAR_FMT
        gc.alignment = CENTER
        gc.fill = PatternFill("solid", fgColor=C_TOTAL)
        gc.border = BORDER_TOTAL
        dgc = ws.cell(row=r, column=5, value=f"=SUM(E{first}:E{last})")
        dgc.font = F_TOTAL
        dgc.number_format = QTY_FMT
        dgc.alignment = CENTER
        dgc.fill = PatternFill("solid", fgColor=C_TOTAL)
        dgc.border = BORDER_TOTAL
        ws.cell(row=r, column=2).fill = PatternFill("solid", fgColor=C_TOTAL)
        ws.cell(row=r, column=2).border = BORDER_TOTAL
        grand_days = r

        r += 2
        self._summary_section(ws, r, "المدة الزمنية التقديرية للمشروع")
        r += 1
        timeline = [
            ("إجمالي أيام العمل", f"=E{grand_days}", QTY_FMT),
            (f"بالأسابيع (أسبوع = {self.week_days} يوم)", f"=E{grand_days}/{self.week_days}", "#,##0.0"),
            (f"بالأشهر (شهر = {self.month_days} يوم)", f"=E{grand_days}/{self.month_days}", "#,##0.0"),
        ]
        for label, val, fmt in timeline:
            lc = ws.cell(row=r, column=3, value=label)
            lc.alignment = RIGHT
            lc.font = F_CELL
            lc.border = BORDER
            vc = ws.cell(row=r, column=4, value=val)
            vc.number_format = fmt
            vc.alignment = CENTER
            vc.font = F_CELL
            vc.border = BORDER
            r += 1
        ws.freeze_panes = "A7"

    def _summary_section(self, ws, row, text):
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        c = ws.cell(row=row, column=2, value=text)
        c.font = F_SECTION
        c.alignment = RIGHT
        c.fill = PatternFill("solid", fgColor=C_BRAND)

    # -- helpers -------------------------------------------------------------
    def _place_image(self, ws, item, data_dir, anchor):
        rel = item.get("image")
        if not rel:
            return
        raw = None
        try:
            from store import read_image_bytes
            raw = read_image_bytes(rel)
        except Exception:
            raw = None
        if raw is None:
            path = rel if os.path.isabs(rel) else os.path.join(data_dir, rel)
            if not os.path.exists(path):
                return
            with open(path, "rb") as f:
                raw = f.read()
        try:
            img = XLImage(io.BytesIO(raw))
            max_w, max_h = 300, 150
            if img.width and img.height:
                ratio = min(max_w / img.width, max_h / img.height, 1.0)
                img.width = int(img.width * ratio)
                img.height = int(img.height * ratio)
            ws.add_image(img, anchor)
        except Exception:
            pass

    @staticmethod
    def _num(value, default=0.0):
        try:
            if value in (None, ""):
                return default
            return float(value)
        except (TypeError, ValueError):
            return default
